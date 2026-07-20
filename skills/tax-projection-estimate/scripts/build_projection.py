#!/usr/bin/env python3
"""
build_projection.py — Reusable scaffolding builder for the firm-stack
`tax-projection-estimate` skill (Form 1065 / Form 1120S pass-through projections).

WHAT THIS IS
------------
A styling + formula-wiring helper, NOT a one-client generator. The projection
content is data-dependent (it comes from the client's trial balance and prior-year
return), so this module gives you:

  (a) Palette / font / number-format / column-width constants that match
      references/workbook-build.md and references/account-tagging.md EXACTLY.
  (b) Helpers to build:
        - the tagged trial-balance sheet, with the Tax Group tag column colored
          by category (the optional visual-scanning fills), and
        - the projection sheet whose income/expense lines are wired with real
          `=SUMIF(...)` formulas that reference the TB tag column by the label
          cell in the same row (never a hard-coded sum, never a hard-coded tag
          inside the formula), plus a per-owner K-1 allocation grid.
  (c) Enforcement of the documented failure mode: on the projection sheet,
      DATA ROWS GET NO FILL. Only structural rows (section headers, column
      headers, totals, subtotals, flagged items) are ever filled. Data rows go
      through a writer that cannot apply a fill, and `audit_no_data_fill()`
      verifies it after the fact.

The only fill that ever lands on a data-content row is the Tax Group *tag column*
on the TB sheet — that is the documented, intentional exception (tags are a
categorization aid, not a data value). The account name and amount cells on the
TB are still left white.

Run `python build_projection.py --demo` to build a small 2-shareholder 1120S
example that proves the SUMIF wiring and the no-fill-on-data-rows rule.

openpyxl only. Tested against openpyxl 3.1.5.
"""

from __future__ import annotations

import argparse
import os
import tempfile
from dataclasses import dataclass, field
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ===========================================================================
# 1. CONSTANTS — must match references/workbook-build.md + account-tagging.md
# ===========================================================================

# --- Section (projection sheet) palette. Exactly 5 structural fills. ---------
# From workbook-build.md "Styling Rules". Data rows use NONE of these.
FILL_SECTION_HEADER = "1F4E79"   # dark blue  — section headers
FILL_COLUMN_HEADER = "D6E4F0"    # soft blue  — column headers
FILL_TOTAL = "E2EFDA"            # soft green — total rows
FILL_SUBTOTAL = "F2F2F2"         # light gray — subtotal rows
FILL_FLAGGED = "FFF2CC"          # soft yellow — flagged items

# --- Trial Balance tag-column fills (optional visual scanning) --------------
# From workbook-build.md "Trial Balance tag column fills" table. Keyed by the
# tag *category*, not the individual tag.
TB_TAG_FILLS = {
    "income": "E2EFDA",          # light green
    "expense": "FDE9D9",         # light orange
    "cogs": "D6E4F0",            # light blue
    "bs": "F2F2F2",              # light gray
    "k1": "E8D5F5",              # light purple
    "nondeductible": "FFC7CE",   # light red
    "capital": "FFF2CC",         # light yellow (capital / reclassify)
}

# --- Fonts (from the Styling Rules table) -----------------------------------
FONT_SECTION_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_COLUMN_HEADER = Font(bold=True, size=10)
FONT_TOTAL = Font(bold=True, size=10)
FONT_SUBTOTAL = Font(bold=True, size=10)
FONT_FLAGGED = Font(bold=True, size=10)
FONT_DATA = Font(bold=False, size=10)                       # normal 10pt
FONT_NEGATIVE = Font(color="C00000", size=10)               # losses / negatives
FONT_NOTES = Font(color="666666", italic=True, size=9)      # notes column
FONT_NONTAXABLE_K1 = Font(color="666666", size=10)          # non-taxable K-1 items

# --- Number formats ---------------------------------------------------------
FMT_PLAIN = "#,##0"                 # non-negative amounts
FMT_SIGNED = "#,##0;(#,##0)"        # amounts that can be negative

# --- Column widths (from workbook-build.md: B=44, C-F=15-17, Notes=45) ------
WIDTH_DESC = 44          # column B (description / label — this IS the tag)
WIDTH_AMOUNT = 16        # columns C-F (15-17 range)
WIDTH_NOTES = 45         # notes column

# --- Borders (total rows: top=medium, bottom=double; subtotal: bottom=thin) -
BORDER_TOTAL = Border(top=Side(style="medium"), bottom=Side(style="double"))
BORDER_SUBTOTAL = Border(bottom=Side(style="thin"))

ALIGN_RIGHT = Alignment(horizontal="right")
ALIGN_LEFT = Alignment(horizontal="left")


def _solid(hex_color: str) -> PatternFill:
    """Solid PatternFill from a 6-digit hex string."""
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# ===========================================================================
# 2. TAG -> CATEGORY resolution (for the TB tag-column fills)
# ===========================================================================
# Tags come from references/account-tagging.md. We map each tag to one of the
# seven fill categories in TB_TAG_FILLS. Matching is by prefix/keyword so the
# open-ended tags (CAPITAL-*, RECLASSIFY-*, K1-*, BS-*, NONDEDUCTIBLE-*) resolve
# without an exhaustive list.

_INCOME_TAGS = {
    "gross receipts", "sales returns", "rental income", "brokerage income",
    "interest income", "other income", "gain/loss on sale",
}
_COGS_TAGS = {"cogs"}
_EXPENSE_TAGS = {
    "salaries & wages", "officer compensation", "commissions",
    "repairs & maintenance", "rents", "taxes & licenses", "interest expense",
    "advertising", "insurance", "hoa expenses", "property taxes", "utilities",
    "professional fees", "office expenses", "dues & subscriptions",
    "employee benefits", "pension/retirement", "bad debts", "depreciation",
    "travel", "meals", "vehicle expenses", "contract labor", "other deductions",
}


def tag_category(tag: str) -> Optional[str]:
    """Resolve a tag string to a TB_TAG_FILLS category key, or None if unknown.

    Returns one of: income, cogs, expense, k1, nondeductible, capital, bs.
    """
    if not tag:
        return None
    t = tag.strip()
    low = t.lower()

    # Prefix-based (open-ended) tags first.
    if low.startswith("k1-"):
        return "k1"
    if low.startswith("nondeductible"):
        return "nondeductible"
    if low.startswith("capital-") or low.startswith("reclassify-"):
        return "capital"
    if low.startswith("bs-"):
        return "bs"

    # Exact-set membership.
    if low in _INCOME_TAGS:
        return "income"
    if low in _COGS_TAGS:
        return "cogs"
    if low in _EXPENSE_TAGS:
        return "expense"
    return None


def tag_fill(tag: str) -> Optional[PatternFill]:
    """PatternFill for a tag's category, or None if the category is unknown."""
    cat = tag_category(tag)
    if cat is None:
        return None
    return _solid(TB_TAG_FILLS[cat])


# ===========================================================================
# 3. TRIAL BALANCE SHEET BUILDER
# ===========================================================================

@dataclass
class TBAccount:
    """One trial-balance row. `net` is DR (CR) — debits positive, credits
    negative — matching the skill's normalized column D."""
    name: str
    net: float
    tag: str
    debit: Optional[float] = None
    credit: Optional[float] = None

    def dr(self) -> float:
        if self.debit is not None:
            return self.debit
        return self.net if self.net > 0 else 0.0

    def cr(self) -> float:
        if self.credit is not None:
            return self.credit
        return -self.net if self.net < 0 else 0.0


# Normalized TB layout from SKILL.md Phase 1 + tag column from Phase 2:
#   A: Account name | B: Debit | C: Credit | D: DR (CR) net | E: Tax Group
TB_COL_NAME = 1
TB_COL_DEBIT = 2
TB_COL_CREDIT = 3
TB_COL_NET = 4
TB_COL_TAG = 5
TB_HEADER_ROW = 1
TB_DATA_START = 2


def build_tb_sheet(ws: Worksheet, accounts: list[TBAccount],
                   color_tags: bool = True) -> str:
    """Write the tagged trial-balance sheet.

    - Header row (row 1) uses the column-header fill.
    - Data rows carry NO fill on the name/amount cells (columns A-D).
    - The Tax Group column (E) is colored by tag category when color_tags=True
      (the documented, intentional exception — a categorization aid, not data).
    - A dropdown data-validation is attached to the tag column for consistency.

    Returns the sheet title (for use in SUMIF references).
    """
    headers = ["Account", "Debit", "Credit", "DR (CR)", "Tax Group"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=TB_HEADER_ROW, column=col, value=text)
        c.fill = _solid(FILL_COLUMN_HEADER)
        c.font = FONT_COLUMN_HEADER
        c.alignment = ALIGN_RIGHT if col in (TB_COL_DEBIT, TB_COL_CREDIT, TB_COL_NET) else ALIGN_LEFT

    for i, acct in enumerate(accounts):
        r = TB_DATA_START + i
        # Name — data cell, NO fill.
        ws.cell(row=r, column=TB_COL_NAME, value=acct.name).font = FONT_DATA
        # Debit / Credit / Net — data cells, NO fill.
        for col, val in ((TB_COL_DEBIT, acct.dr()),
                         (TB_COL_CREDIT, acct.cr()),
                         (TB_COL_NET, acct.net)):
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = FMT_SIGNED
            cell.alignment = ALIGN_RIGHT
            cell.font = FONT_DATA
        # Tag column — the one allowed data-row fill.
        tag_cell = ws.cell(row=r, column=TB_COL_TAG, value=acct.tag)
        tag_cell.font = FONT_DATA
        if color_tags:
            fill = tag_fill(acct.tag)
            if fill is not None:
                tag_cell.fill = fill

    # Column widths.
    ws.column_dimensions[get_column_letter(TB_COL_NAME)].width = WIDTH_DESC
    for col in (TB_COL_DEBIT, TB_COL_CREDIT, TB_COL_NET):
        ws.column_dimensions[get_column_letter(col)].width = WIDTH_AMOUNT
    ws.column_dimensions[get_column_letter(TB_COL_TAG)].width = 26

    # Dropdown for tag consistency (Phase 2). List a representative tag set.
    _add_tag_dropdown(ws, len(accounts))
    return ws.title


def _add_tag_dropdown(ws: Worksheet, n_accounts: int) -> None:
    """Attach an Excel data-validation dropdown to the tag column."""
    tags = [
        "Gross Receipts", "Sales Returns", "Rental Income", "Interest Income",
        "Other Income", "Gain/Loss on Sale", "COGS", "Officer Compensation",
        "Salaries & Wages", "Repairs & Maintenance", "Rents", "Taxes & Licenses",
        "Interest Expense", "Advertising", "Insurance", "Depreciation", "Meals",
        "Other Deductions", "K1-Interest", "K1-Dividends", "K1-179",
        "NONDEDUCTIBLE-Meals50", "NONDEDUCTIBLE-Entertainment",
        "BS-Cash", "BS-FixedAsset", "BS-AccumDepr", "BS-Equity", "BS-Loans",
    ]
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(tags), allow_blank=False)
    last = TB_DATA_START + max(n_accounts, 1) - 1
    col = get_column_letter(TB_COL_TAG)
    dv.add(f"{col}{TB_DATA_START}:{col}{last}")
    ws.add_data_validation(dv)


# ===========================================================================
# 4. PROJECTION SHEET — STRUCTURAL vs DATA ROW WRITERS
# ===========================================================================
# The whole no-fill discipline is architectural: data content only ever goes
# through `write_data_row` (which cannot set a fill). Fills are only reachable
# through the structural writers below. `data_rows` on the ProjectionWriter
# records every data row so `audit_no_data_fill` can verify the invariant.

# Projection column layout (workbook-build.md): B is the label (the tag).
PROJ_COL_DESC = 2      # B: Description / label (IS the tag for SUMIF lines)
PROJ_COL_PRIOR = 3     # C: Prior-year actual (hardcoded input)
PROJ_COL_CURR = 4      # D: Current-year estimated (SUMIF formula)
PROJ_COL_VAR = 5       # E: Variance (=D-C)
PROJ_COL_NOTES = 6     # F: Notes


@dataclass
class ProjectionWriter:
    """Cursor + row bookkeeping for building a projection sheet."""
    ws: Worksheet
    tb_sheet: str
    row: int = 1
    data_rows: list[int] = field(default_factory=list)   # rows that MUST stay unfilled
    fill_rows: list[int] = field(default_factory=list)    # structural rows (allowed fills)
    rows: dict = field(default_factory=dict)              # named row lookups for formulas

    # -- structural rows (fills allowed) ------------------------------------
    def section_header(self, text: str) -> int:
        r = self.row
        c = self.ws.cell(row=r, column=PROJ_COL_DESC, value=text)
        c.fill = _solid(FILL_SECTION_HEADER)
        c.font = FONT_SECTION_HEADER
        # Extend the fill across the used columns for a clean band.
        for col in range(PROJ_COL_PRIOR, PROJ_COL_NOTES + 1):
            self.ws.cell(row=r, column=col).fill = _solid(FILL_SECTION_HEADER)
        self.fill_rows.append(r)
        self.row += 1
        return r

    def column_headers(self, labels: dict) -> int:
        """labels: {col_index: text}."""
        r = self.row
        for col, text in labels.items():
            c = self.ws.cell(row=r, column=col, value=text)
            c.fill = _solid(FILL_COLUMN_HEADER)
            c.font = FONT_COLUMN_HEADER
            c.alignment = ALIGN_RIGHT if col >= PROJ_COL_PRIOR and col != PROJ_COL_NOTES else ALIGN_LEFT
        self.fill_rows.append(r)
        self.row += 1
        return r

    def total_row(self, label: str, formula_cols: dict, note: str = "") -> int:
        """Total row: soft-green fill, bold, top=medium/bottom=double border.
        formula_cols: {col_index: formula_or_value}."""
        r = self.row
        lc = self.ws.cell(row=r, column=PROJ_COL_DESC, value=label)
        lc.fill = _solid(FILL_TOTAL)
        lc.font = FONT_TOTAL
        lc.border = BORDER_TOTAL
        for col in range(PROJ_COL_PRIOR, PROJ_COL_NOTES + 1):
            cell = self.ws.cell(row=r, column=col, value=formula_cols.get(col))
            cell.fill = _solid(FILL_TOTAL)
            cell.font = FONT_TOTAL
            cell.border = BORDER_TOTAL
            if col != PROJ_COL_NOTES:
                cell.number_format = FMT_SIGNED
                cell.alignment = ALIGN_RIGHT
        if note:
            self.ws.cell(row=r, column=PROJ_COL_NOTES, value=note).font = FONT_NOTES
        self.fill_rows.append(r)
        self.row += 1
        return r

    def subtotal_row(self, label: str, formula_cols: dict) -> int:
        """Subtotal row: light-gray fill, bold, bottom=thin border."""
        r = self.row
        lc = self.ws.cell(row=r, column=PROJ_COL_DESC, value=label)
        lc.fill = _solid(FILL_SUBTOTAL)
        lc.font = FONT_SUBTOTAL
        lc.border = BORDER_SUBTOTAL
        for col in range(PROJ_COL_PRIOR, PROJ_COL_NOTES + 1):
            cell = self.ws.cell(row=r, column=col, value=formula_cols.get(col))
            cell.fill = _solid(FILL_SUBTOTAL)
            cell.font = FONT_SUBTOTAL
            cell.border = BORDER_SUBTOTAL
            if col != PROJ_COL_NOTES:
                cell.number_format = FMT_SIGNED
                cell.alignment = ALIGN_RIGHT
        self.fill_rows.append(r)
        self.row += 1
        return r

    def flagged_row(self, label: str, values: dict, note: str = "") -> int:
        """Flagged item: soft-yellow fill, bold. Use for open items / warnings."""
        r = self.row
        lc = self.ws.cell(row=r, column=PROJ_COL_DESC, value=label)
        lc.fill = _solid(FILL_FLAGGED)
        lc.font = FONT_FLAGGED
        for col in range(PROJ_COL_PRIOR, PROJ_COL_NOTES + 1):
            cell = self.ws.cell(row=r, column=col, value=values.get(col))
            cell.fill = _solid(FILL_FLAGGED)
            cell.font = FONT_FLAGGED
            if col != PROJ_COL_NOTES:
                cell.number_format = FMT_SIGNED
                cell.alignment = ALIGN_RIGHT
        if note:
            self.ws.cell(row=r, column=PROJ_COL_NOTES, value=note).font = FONT_FLAGGED
        self.fill_rows.append(r)
        self.row += 1
        return r

    # -- data rows (NO fill — the whole point) ------------------------------
    def data_row(self, label: str, prior=None, curr_formula=None,
                 note: str = "", nontaxable: bool = False) -> int:
        """A data line. NEVER sets a fill.

        `label` goes in column B and IS the SUMIF criteria (the tag). `prior`
        is a hardcoded prior-year input. `curr_formula` is the column-D formula
        (typically a SUMIF referencing B{r}). Variance = D - C is auto-wired
        when both exist.
        """
        r = self.row
        # Label — no fill.
        lc = self.ws.cell(row=r, column=PROJ_COL_DESC, value=label)
        lc.font = FONT_NONTAXABLE_K1 if nontaxable else FONT_DATA
        # Prior year (hardcoded input) — no fill.
        if prior is not None:
            pc = self.ws.cell(row=r, column=PROJ_COL_PRIOR, value=prior)
            pc.number_format = FMT_SIGNED
            pc.alignment = ALIGN_RIGHT
            pc.font = FONT_NONTAXABLE_K1 if nontaxable else FONT_DATA
        # Current year (formula) — no fill.
        if curr_formula is not None:
            cc = self.ws.cell(row=r, column=PROJ_COL_CURR, value=curr_formula)
            cc.number_format = FMT_SIGNED
            cc.alignment = ALIGN_RIGHT
            cc.font = FONT_NONTAXABLE_K1 if nontaxable else FONT_DATA
        # Variance = D - C — no fill.
        if prior is not None and curr_formula is not None:
            vc = self.ws.cell(row=r, column=PROJ_COL_VAR,
                              value=f"=D{r}-C{r}")
            vc.number_format = FMT_SIGNED
            vc.alignment = ALIGN_RIGHT
            vc.font = FONT_DATA
        if note:
            self.ws.cell(row=r, column=PROJ_COL_NOTES, value=note).font = FONT_NOTES
        self.data_rows.append(r)
        self.row += 1
        return r

    def blank(self) -> None:
        self.row += 1


# --- SUMIF formula builders (never hard-code a sum; never hard-code a tag) --

def income_sumif(tb_sheet: str, label_row: int) -> str:
    """Income line: credits are negative on the TB, so negate the SUMIF.
    Criteria is the label cell B{row} — the tag lives in column B, never in
    the formula text."""
    return (f"=-SUMIF('{tb_sheet}'!$E:$E,B{label_row},"
            f"'{tb_sheet}'!$D:$D)")


def expense_sumif(tb_sheet: str, label_row: int) -> str:
    """Expense line: debits are positive on the TB, so SUMIF directly."""
    return (f"=SUMIF('{tb_sheet}'!$E:$E,B{label_row},"
            f"'{tb_sheet}'!$D:$D)")


# ===========================================================================
# 5. K-1 ALLOCATION GRID (per-owner) — the primary deliverable
# ===========================================================================
# Columns: B: K-1 Line | C: Description | D: Total | E..: Owner (pct)
# Owner cells are formulas referencing the Total column (=D{r}*pct).

@dataclass
class Owner:
    name: str
    pct: float   # e.g. 0.5 for 50%


def build_k1_grid(pw: ProjectionWriter, owners: list[Owner],
                  lines: list[dict]) -> None:
    """Append a K-1 allocation grid to the projection.

    `lines` items: {"k1_line": str, "desc": str, "total": <formula|value>,
                    "nontaxable": bool (optional)}.
    Owner columns are wired as =D{r}*pct so they cascade from the Total column.
    """
    pw.section_header("K-1 ALLOCATION BY LINE ITEM")

    # Column headers: K-1 Line | Description | Total | Owner1 (xx%) | ...
    owner_col0 = PROJ_COL_CURR + 1  # E is first owner column
    headers = {PROJ_COL_DESC: "K-1 Line", PROJ_COL_PRIOR: "Description",
               PROJ_COL_CURR: "Total"}
    for i, o in enumerate(owners):
        headers[owner_col0 + i] = f"{o.name} ({o.pct*100:.0f}%)"
    hr = pw.column_headers(headers)

    taxable_line_rows: list[int] = []
    for ln in lines:
        r = pw.row
        nontax = ln.get("nontaxable", False)
        fnt = FONT_NONTAXABLE_K1 if nontax else FONT_DATA
        # K-1 line number (B) — data cell, no fill.
        pw.ws.cell(row=r, column=PROJ_COL_DESC, value=ln["k1_line"]).font = fnt
        # Description (C) — data cell, no fill.
        pw.ws.cell(row=r, column=PROJ_COL_PRIOR, value=ln["desc"]).font = fnt
        # Total (D) — formula/value referencing the sections above.
        tc = pw.ws.cell(row=r, column=PROJ_COL_CURR, value=ln["total"])
        tc.number_format = FMT_SIGNED
        tc.alignment = ALIGN_RIGHT
        tc.font = fnt
        # Owner columns — =D{r}*pct, cascading from the Total.
        for i, o in enumerate(owners):
            col = owner_col0 + i
            oc = pw.ws.cell(row=r, column=col,
                            value=f"=D{r}*{o.pct}")
            oc.number_format = FMT_SIGNED
            oc.alignment = ALIGN_RIGHT
            oc.font = fnt
        pw.data_rows.append(r)
        if not nontax:
            taxable_line_rows.append(r)
        pw.row += 1

    # Est. taxable K-1 income total (SUM of taxable line rows), per owner.
    if taxable_line_rows:
        total_formula_cols: dict = {}
        # Total column D.
        d_refs = ",".join(f"D{r}" for r in taxable_line_rows)
        total_formula_cols[PROJ_COL_CURR] = f"=SUM({d_refs})"
        # Each owner column sums its own taxable cells.
        for i, o in enumerate(owners):
            col = owner_col0 + i
            letter = get_column_letter(col)
            refs = ",".join(f"{letter}{r}" for r in taxable_line_rows)
            total_formula_cols[col] = f"=SUM({refs})"
        # Reuse total_row but it only fills B..F; extend to owner cols manually.
        tr = pw.total_row("Est. Taxable K-1 Income", total_formula_cols,
                          note="Excl. distributions, nondeductible, tax-exempt")
        # total_row already styled B..F; ensure any owner cols beyond F styled.
        for i in range(len(owners)):
            col = owner_col0 + i
            if col > PROJ_COL_NOTES:
                cell = pw.ws.cell(row=tr, column=col,
                                  value=total_formula_cols[col])
                cell.fill = _solid(FILL_TOTAL)
                cell.font = FONT_TOTAL
                cell.border = BORDER_TOTAL
                cell.number_format = FMT_SIGNED
                cell.alignment = ALIGN_RIGHT


# ===========================================================================
# 6. AUDIT — enforce "data rows have NO fill"
# ===========================================================================

def audit_no_data_fill(ws: Worksheet, data_rows: list[int]) -> list[str]:
    """Return a list of violations: any cell on a recorded data row that has a
    solid fill. Empty list == the invariant holds. This is the automated guard
    against the documented failure mode (subagents filling data rows)."""
    violations = []
    for r in data_rows:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            fill = cell.fill
            if fill is not None and fill.fill_type == "solid":
                violations.append(
                    f"row {r}, col {get_column_letter(col)} has fill "
                    f"{fill.fgColor.rgb}")
    return violations


# ===========================================================================
# 7. DEMO — small 2-shareholder 1120S; proves SUMIF wiring + no-fill rule
# ===========================================================================

def build_demo(out_path: str) -> str:
    """Build a minimal but complete 1120S projection skeleton and save it.

    Demonstrates: tagged TB with colored tag column, income (-SUMIF) and
    expense (SUMIF) lines wired to the tag column, ordinary business income
    total, and a 2-shareholder K-1 grid whose owner columns cascade from the
    Total column. Returns out_path.
    """
    wb = Workbook()

    # --- Projection sheet must be FIRST (SKILL.md Phase 4). Create it now,
    #     but we fill the TB first so we know its title for SUMIF refs. ------
    proj_ws = wb.active
    proj_ws.title = "2026 Tax Projection"
    tb_ws = wb.create_sheet("Trial Balance")

    # --- Tagged trial balance (net = DR (CR): credits negative) ------------
    accounts = [
        TBAccount("Sales - Services", net=-1_200_000, tag="Gross Receipts"),
        TBAccount("Interest Income", net=-3_500, tag="K1-Interest"),
        TBAccount("Cost of Services", net=430_000, tag="COGS"),
        TBAccount("Officer Salary - Shareholders", net=180_000, tag="Officer Compensation"),
        TBAccount("Staff Wages", net=210_000, tag="Salaries & Wages"),
        TBAccount("Repairs", net=14_500, tag="Repairs & Maintenance"),
        TBAccount("Meals (100% on books)", net=9_000, tag="Meals"),
        TBAccount("Meals 50% addback", net=4_500, tag="NONDEDUCTIBLE-Meals50"),
        TBAccount("Book Depreciation", net=52_000, tag="Depreciation"),
        TBAccount("Operating Cash", net=140_000, tag="BS-Cash"),
        TBAccount("Shareholder Distributions", net=90_000, tag="BS-Equity"),
        TBAccount("Common Stock / APIC", net=-16_000, tag="BS-Equity"),
    ]
    tb_title = build_tb_sheet(tb_ws, accounts, color_tags=True)

    # --- Projection sheet ---------------------------------------------------
    pw = ProjectionWriter(proj_ws, tb_sheet=tb_title, row=1)

    # HEADER BLOCK (data rows — no fill).
    pw.data_row("Entity: Demo Services Inc.", note="Form 1120S — S-corporation")
    pw.data_row("Tax Year: 2026", note="Preliminary estimate")
    pw.blank()

    # INCOME.
    pw.section_header("INCOME")
    pw.column_headers({PROJ_COL_DESC: "Description",
                       PROJ_COL_PRIOR: "2025 Actual",
                       PROJ_COL_CURR: "2026 Estimated",
                       PROJ_COL_VAR: "Variance",
                       PROJ_COL_NOTES: "Notes"})
    # Label cell B{r} == tag. Formula pulls -SUMIF from the tag column.
    r_gr = pw.row
    pw.data_row("Gross Receipts", prior=1_050_000,
                curr_formula=income_sumif(tb_title, r_gr),
                note="Line 1a")
    r_cogs = pw.row
    pw.data_row("COGS", prior=380_000,
                curr_formula=expense_sumif(tb_title, r_cogs),
                note="Line 2")
    gross_profit = pw.total_row(
        "Gross Profit",
        {PROJ_COL_PRIOR: f"=C{r_gr}-C{r_cogs}",
         PROJ_COL_CURR: f"=D{r_gr}-D{r_cogs}",
         PROJ_COL_VAR: None},
        note="Line 3")
    pw.rows["gross_profit"] = gross_profit
    pw.blank()

    # DEDUCTIONS.
    pw.section_header("DEDUCTIONS / OPERATING EXPENSES")
    pw.column_headers({PROJ_COL_DESC: "Description",
                       PROJ_COL_PRIOR: "2025 Actual",
                       PROJ_COL_CURR: "2026 Estimated",
                       PROJ_COL_VAR: "Variance",
                       PROJ_COL_NOTES: "Notes"})
    exp_first = pw.row
    r_off = pw.row
    pw.data_row("Officer Compensation", prior=170_000,
                curr_formula=expense_sumif(tb_title, r_off), note="Line 7")
    r_sal = pw.row
    pw.data_row("Salaries & Wages", prior=195_000,
                curr_formula=expense_sumif(tb_title, r_sal), note="Line 8")
    r_rep = pw.row
    pw.data_row("Repairs & Maintenance", prior=12_000,
                curr_formula=expense_sumif(tb_title, r_rep), note="Line 9")
    r_dep = pw.row
    pw.data_row("Depreciation", prior=48_000,
                curr_formula=expense_sumif(tb_title, r_dep), note="Line 14 (book)")
    exp_last = pw.row - 1
    total_ded = pw.total_row(
        "Total Deductions",
        {PROJ_COL_PRIOR: f"=SUM(C{exp_first}:C{exp_last})",
         PROJ_COL_CURR: f"=SUM(D{exp_first}:D{exp_last})",
         PROJ_COL_VAR: None})
    pw.rows["total_ded"] = total_ded
    pw.blank()

    # NET / ORDINARY BUSINESS INCOME.
    pw.section_header("NET ORDINARY BUSINESS INCOME")
    obi = pw.total_row(
        "Ordinary Business Income (Line 21)",
        {PROJ_COL_PRIOR: f"=C{gross_profit}-C{total_ded}",
         PROJ_COL_CURR: f"=D{gross_profit}-D{total_ded}",
         PROJ_COL_VAR: None})
    pw.rows["obi"] = obi
    pw.blank()

    # K-1 ALLOCATION GRID — 2 shareholders, 60/40.
    owners = [Owner("Alex Stone", 0.60), Owner("Blair Kim", 0.40)]
    k1_lines = [
        {"k1_line": "Box 1", "desc": "Ordinary business income",
         "total": f"=D{obi}"},
        {"k1_line": "Box 4", "desc": "Interest income",
         "total": income_sumif_for_grid(tb_title, "K1-Interest")},
        {"k1_line": "Box 16C", "desc": "Nondeductible expenses",
         "total": expense_sumif_for_grid(tb_title, "NONDEDUCTIBLE-Meals50"),
         "nontaxable": True},
        {"k1_line": "Box 16D", "desc": "Distributions",
         "total": expense_sumif_for_grid(tb_title, "BS-Equity"),
         "nontaxable": True},
    ]
    build_k1_grid(pw, owners, k1_lines)
    pw.blank()

    # A deliberately flagged open item, to exercise the flagged-row style.
    pw.section_header("NOTES & OPEN ITEMS")
    pw.flagged_row("Book depreciation used; confirm tax dep from prior return",
                   {}, note="Open item")

    # Column widths on the projection sheet.
    proj_ws.column_dimensions[get_column_letter(PROJ_COL_DESC)].width = WIDTH_DESC
    for col in (PROJ_COL_PRIOR, PROJ_COL_CURR, PROJ_COL_VAR):
        proj_ws.column_dimensions[get_column_letter(col)].width = WIDTH_AMOUNT
    proj_ws.column_dimensions[get_column_letter(PROJ_COL_NOTES)].width = WIDTH_NOTES
    # Owner columns.
    for i in range(len(owners)):
        proj_ws.column_dimensions[get_column_letter(PROJ_COL_CURR + 1 + i)].width = WIDTH_AMOUNT

    # Enforce the no-fill-on-data-rows rule before saving.
    violations = audit_no_data_fill(proj_ws, pw.data_rows)
    if violations:
        raise AssertionError("Data-row fill violations: " + "; ".join(violations))

    wb.save(out_path)
    return out_path


def income_sumif_for_grid(tb_sheet: str, tag: str) -> str:
    """K-1 grid Total cell for an income tag. The grid's Total column (D) has no
    same-row label to reference, so the tag is passed explicitly here — but it
    is still a real SUMIF against the TB, never a hard-coded sum."""
    return f'=-SUMIF(\'{tb_sheet}\'!$E:$E,"{tag}",\'{tb_sheet}\'!$D:$D)'


def expense_sumif_for_grid(tb_sheet: str, tag: str) -> str:
    """K-1 grid Total cell for an expense/BS tag (positive on TB)."""
    return f'=SUMIF(\'{tb_sheet}\'!$E:$E,"{tag}",\'{tb_sheet}\'!$D:$D)'


# ===========================================================================
# 8. CLI
# ===========================================================================

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--demo", action="store_true",
                    help="Build a 2-shareholder 1120S demo workbook.")
    ap.add_argument("--out", default=None,
                    help="Output path (default: a temp file, NOT the skill dir).")
    args = ap.parse_args()

    if args.demo:
        out = args.out or os.path.join(tempfile.gettempdir(),
                                       "tax_projection_demo.xlsx")
        path = build_demo(out)
        print(f"Demo workbook written to: {path}")
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
