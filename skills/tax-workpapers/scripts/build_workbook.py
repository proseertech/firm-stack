#!/usr/bin/env python3
"""
build_workbook.py — reusable scaffolding builder for firm-stack tax-workpapers.

Purpose
-------
The *content* of a tax-workpapers workbook is data-dependent (transcribed from a
client's 1099s and K-1s), so this module does NOT invent data. Its job is to
remove the repeated *scaffolding* work: creating styled sheets, header rows,
blue-font input rows, and black-font `=SUM()` total rows that follow the firm's
formatting standards and the financial-model color convention — so every workbook
starts consistent and every total is a real formula, never a hard-coded number.

Everything here mirrors, exactly, the standards in
`references/workbook-structure.md`:

  * Headers        — white bold text on blue fill (#4472C4)
  * Number format  — #,##0.00;(#,##0.00)  (parentheses for negatives)
  * Money columns  — minimum width 14
  * Freeze panes   — frozen below the header row
  * Borders        — thin borders on all data cells
  * Alt row fill   — light green (#E2EFDA) on even data rows
  * Net total row  — bold, double-underline top border
  * Page setup     — landscape, fit to width

Color convention (financial model):
  * BLUE  font (#0000FF) — hard-coded inputs transcribed from source documents
  * BLACK font (#000000) — all formulas and calculations (SUM, =C-D-E, x-sheet)
  * YELLOW fill (#FFFF00) — placeholders where source data is missing/estimated

Dependencies: openpyxl only (tested against 3.1.5). Python 3.

Usage
-----
As a library:
    from build_workbook import (
        new_workbook, add_sheet, write_header, write_data_row,
        write_sum_total_row, finalize_sheet, save,
    )

As a demo (proves the API end to end):
    python build_workbook.py --demo [--out /tmp/demo_workpapers.xlsx]
"""

from __future__ import annotations

import argparse
import os
import string
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Palette / font / format constants — match references/workbook-structure.md
# ---------------------------------------------------------------------------

# Fills
HEADER_FILL_RGB = "4472C4"          # blue header fill
ALT_ROW_FILL_RGB = "E2EFDA"         # light green alternating row fill
PLACEHOLDER_FILL_RGB = "FFFF00"     # yellow placeholder fill

# Font colors (the blue-input / black-formula convention)
INPUT_FONT_RGB = "0000FF"           # BLUE  — hard-coded inputs from source docs
FORMULA_FONT_RGB = "000000"         # BLACK — formulas & calculations
HEADER_FONT_RGB = "FFFFFF"          # WHITE — header text

# Number format (parentheses for negatives)
MONEY_FORMAT = "#,##0.00;(#,##0.00)"

# Column sizing
MIN_MONEY_WIDTH = 14
DEFAULT_LABEL_WIDTH = 28

# --- Reusable style objects -------------------------------------------------

HEADER_FILL = PatternFill(start_color=HEADER_FILL_RGB, end_color=HEADER_FILL_RGB, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=ALT_ROW_FILL_RGB, end_color=ALT_ROW_FILL_RGB, fill_type="solid")
PLACEHOLDER_FILL = PatternFill(start_color=PLACEHOLDER_FILL_RGB, end_color=PLACEHOLDER_FILL_RGB, fill_type="solid")

HEADER_FONT = Font(bold=True, color=HEADER_FONT_RGB)
INPUT_FONT = Font(color=INPUT_FONT_RGB)        # blue — typed-in values
FORMULA_FONT = Font(color=FORMULA_FONT_RGB)    # black — computed values
TOTAL_FONT = Font(bold=True, color=FORMULA_FONT_RGB)  # totals are formulas -> black, bold

_THIN = Side(style="thin")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Net total row: bold with a double-underline top border (a "double" side).
_DOUBLE = Side(style="double")
TOTAL_BORDER = Border(left=_THIN, right=_THIN, top=_DOUBLE, bottom=_THIN)

RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left")
CENTER = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def new_workbook() -> Workbook:
    """Return a fresh workbook with the default sheet removed."""
    wb = Workbook()
    # Drop the auto-created "Sheet" so callers only get the tabs they add.
    default = wb.active
    wb.remove(default)
    return wb


def add_sheet(wb: Workbook, title: str):
    """Add a worksheet with firm page setup (landscape, fit to width)."""
    ws = wb.create_sheet(title=title)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def write_header(ws, headers: Sequence[str], row: int = 1):
    """
    Write a styled header row (white bold on blue) and freeze panes just below
    it. Returns the row index that the first data row should occupy.
    """
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    # Freeze below the header row (e.g. header on row 1 -> freeze "A2").
    ws.freeze_panes = f"A{row + 1}"
    return row + 1


def write_data_row(
    ws,
    row: int,
    values: Sequence,
    *,
    money_cols: Iterable[int] = (),
    placeholder_cols: Iterable[int] = (),
    stripe: bool = True,
):
    """
    Write one detail row of transcribed inputs.

    Inputs get BLUE font (the color convention for hard-coded, typed-in values).
    `money_cols` (1-based column indexes) receive the money number format and
    right alignment. `placeholder_cols` get the yellow placeholder fill (missing
    or estimated data). Even data rows get the light-green alternating stripe.

    Note: values that are formula strings (start with "=") are still written as
    given, but this helper is intended for INPUTS. Use write_formula_cell /
    write_sum_total_row for computed cells so they get black font.
    """
    money_cols = set(money_cols)
    placeholder_cols = set(placeholder_cols)
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = INPUT_FONT
        cell.border = THIN_BORDER
        if col_idx in money_cols:
            cell.number_format = MONEY_FORMAT
            cell.alignment = RIGHT
        if stripe and row % 2 == 0:
            cell.fill = ALT_ROW_FILL
        if col_idx in placeholder_cols:
            cell.fill = PLACEHOLDER_FILL
    return row + 1


def write_formula_cell(ws, row: int, col: int, formula: str, *, money: bool = True, bold: bool = False):
    """
    Write a single computed cell as a real formula string with BLACK font.
    Used for gain/(loss) columns like `=C5-D5-E5` and cross-sheet references.
    """
    if not formula.startswith("="):
        raise ValueError(f"formula must start with '=', got: {formula!r}")
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = TOTAL_FONT if bold else FORMULA_FONT
    cell.border = THIN_BORDER
    if money:
        cell.number_format = MONEY_FORMAT
        cell.alignment = RIGHT
    return cell


def write_sum_total_row(
    ws,
    row: int,
    label: str,
    sum_cols: Iterable[int],
    first_data_row: int,
    last_data_row: int,
    *,
    label_col: int = 1,
    net_total: bool = False,
):
    """
    Write a total row whose numeric cells are REAL `=SUM(...)` formulas over the
    detail range — never a hard-coded number. Totals are formulas, so they get
    BLACK bold font per the color convention.

    `sum_cols` are 1-based column indexes to total. `net_total=True` applies the
    bold double-underline top border used for the net total row.
    """
    border = TOTAL_BORDER if net_total else THIN_BORDER

    # Label cell.
    lbl = ws.cell(row=row, column=label_col, value=label)
    lbl.font = TOTAL_FONT
    lbl.border = border

    for col_idx in sum_cols:
        letter = get_column_letter(col_idx)
        formula = f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = TOTAL_FONT          # black + bold: it's a computed total
        cell.border = border
        cell.number_format = MONEY_FORMAT
        cell.alignment = RIGHT
    return row + 1


def autosize_columns(ws, money_cols: Iterable[int] = (), label_width: int = DEFAULT_LABEL_WIDTH):
    """
    Set sensible column widths: money columns get at least MIN_MONEY_WIDTH,
    other columns are fit to their longest cell content (with a floor).
    """
    money_cols = set(money_cols)
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        letter = get_column_letter(col_idx)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        if col_idx in money_cols:
            width = max(MIN_MONEY_WIDTH, longest + 2)
        elif col_idx == 1:
            width = max(label_width, longest + 2)
        else:
            width = max(10, longest + 2)
        ws.column_dimensions[letter].width = width


def finalize_sheet(ws, money_cols: Iterable[int] = ()):
    """Convenience wrapper for post-population steps (currently column sizing)."""
    autosize_columns(ws, money_cols=money_cols)


def save(wb: Workbook, path: str):
    """Save the workbook, creating parent dirs as needed."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Demo — builds a small example workbook to prove the API.
# ---------------------------------------------------------------------------

def build_demo(path: str) -> str:
    """
    Build a two-tab example workbook:
      1. "1099-B" capital gains tab — inputs (blue), a `=C-D-E` wash-sale
         gain/(loss) formula column (black), and a `=SUM()` net total row.
      2. "K-1 (1065)" summary tab — one column per partnership with a black
         `=SUM()` TOTAL column.

    Data values here are illustrative placeholders ONLY — the real skill
    transcribes them from the client's source PDFs.
    """
    wb = new_workbook()

    # --- Tab 1: 1099-B (Capital Gains) --------------------------------------
    ws = add_sheet(wb, "1099-B")
    headers = [
        "Custodian / Account", "Box / Category", "Proceeds", "Cost Basis",
        "Wash Sale Adjustment", "Gain / (Loss)", "Term", "Notes",
    ]
    # Money columns: Proceeds(C=3), Cost Basis(D=4), Wash(E=5), Gain(F=6).
    money_cols = [3, 4, 5, 6]
    first_data_row = write_header(ws, headers)

    # Illustrative detail rows (C, D, E are inputs; F is a formula).
    demo_rows = [
        ("GS ADV — Box A ST Covered", "A / Short-Term", 125000.00, 130000.00, -2000.00, "ST", "Wash sale reported"),
        ("GS ADV — Box A LT Covered", "A / Long-Term", 480000.00, 310000.00, None, "LT", ""),
        ("Fidelity — Box D LT Covered", "D / Long-Term", 92000.00, 88000.00, None, "LT", ""),
    ]

    r = first_data_row
    for account, box, proceeds, basis, wash, term, notes in demo_rows:
        # Write the input columns (blue). Leave Gain col (F) empty here.
        # Wash placeholder highlighting: if wash is None, no cell; the reference
        # says leave E empty when embedded in basis. We just leave it blank.
        write_data_row(
            ws, r,
            [account, box, proceeds, basis, wash, None, term, notes],
            money_cols=money_cols,
        )
        # Gain/(Loss) is ALWAYS a formula (black). Use =C-D-E when a wash sale is
        # present, else =C-D. E is blank -> subtracting it is harmless, but we
        # match the reference exactly and omit E when there is no wash sale.
        if wash is not None:
            gain_formula = f"=C{r}-D{r}-E{r}"
        else:
            gain_formula = f"=C{r}-D{r}"
        write_formula_cell(ws, r, 6, gain_formula)
        r += 1
    last_data_row = r - 1

    # Net total row: real =SUM() formulas, bold + double underline.
    write_sum_total_row(
        ws, r, "TOTAL",
        sum_cols=[3, 4, 5, 6],
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        net_total=True,
    )
    finalize_sheet(ws, money_cols=money_cols)

    # --- Tab 2: K-1 (1065) --------------------------------------------------
    ws2 = add_sheet(wb, "K-1 (1065)")
    # One column per partnership, plus a TOTAL column that SUMs across them.
    k1_headers = ["K-1 Line Item", "Alpha Fund LP", "Beta Partners LP", "TOTAL"]
    k1_first = write_header(ws2, k1_headers)
    k1_money = [2, 3, 4]  # the two partnership columns + TOTAL

    k1_rows = [
        ("Box 1: Ordinary Business Income", 45000.00, 12000.00),
        ("Box 5: Interest Income", 3200.00, 900.00),
        ("Box 6a: Ordinary Dividends", 15000.00, 4100.00),
        ("Box 9a: Net LT Capital Gain (Loss)", 82000.00, -5000.00),
    ]

    rr = k1_first
    for label, alpha, beta in k1_rows:
        # Partnership amounts are inputs (blue).
        write_data_row(ws2, rr, [label, alpha, beta, None], money_cols=k1_money)
        # TOTAL column (D) is a per-row SUM across the partnership columns (black).
        write_formula_cell(ws2, rr, 4, f"=SUM(B{rr}:C{rr})")
        rr += 1
    k1_last = rr - 1

    # Column-total row across every partnership + the TOTAL column.
    write_sum_total_row(
        ws2, rr, "TOTAL",
        sum_cols=[2, 3, 4],
        first_data_row=k1_first,
        last_data_row=k1_last,
        net_total=True,
    )
    finalize_sheet(ws2, money_cols=k1_money)

    return save(wb, path)


def _verify(path: str) -> None:
    """
    Reload the demo workbook and assert the key invariants hold:
      * total cells contain real =SUM(...) formulas (not hard-coded numbers)
      * the gain/(loss) column holds =C-D-E / =C-D formulas
      * inputs are blue, formulas/totals are black, headers are white-on-blue
    """
    wb = load_workbook(path, data_only=False)  # keep formulas
    ws = wb["1099-B"]

    # Header styling.
    h = ws["A1"]
    assert h.font.color.rgb.endswith(HEADER_FONT_RGB), f"header font not white: {h.font.color.rgb}"
    assert h.fill.start_color.rgb.endswith(HEADER_FILL_RGB), f"header fill not blue: {h.fill.start_color.rgb}"

    # A blue input cell (Proceeds on first data row = C2).
    c2 = ws["C2"]
    assert c2.font.color.rgb.endswith(INPUT_FONT_RGB), f"input not blue: {c2.font.color.rgb}"
    assert isinstance(c2.value, (int, float)), f"input should be a number: {c2.value!r}"

    # Gain/(loss) formula column (F2) — black formula with wash sale.
    f2 = ws["F2"]
    assert str(f2.value).startswith("=C2-D2-E2"), f"gain formula wrong: {f2.value!r}"
    assert f2.font.color.rgb.endswith(FORMULA_FONT_RGB), f"formula not black: {f2.font.color.rgb}"

    # Second row has no wash sale -> =C3-D3.
    assert str(ws["F3"].value) == "=C3-D3", f"expected =C3-D3, got {ws['F3'].value!r}"

    # Net total row: find the row whose column A == "TOTAL".
    total_row = next(
        cell.row for col in ws.iter_cols(min_col=1, max_col=1) for cell in col if cell.value == "TOTAL"
    )
    proceeds_total = ws.cell(row=total_row, column=3)
    assert str(proceeds_total.value).startswith("=SUM"), \
        f"total is not a SUM formula: {proceeds_total.value!r}"
    assert proceeds_total.font.bold, "total should be bold"
    assert proceeds_total.border.top.style == "double", "net total should have double top border"

    # K-1 tab TOTAL column is a per-row SUM.
    ws2 = wb["K-1 (1065)"]
    assert str(ws2["D2"].value).startswith("=SUM(B2:C2"), f"K-1 total wrong: {ws2['D2'].value!r}"

    print("VERIFY OK:")
    print(f"  1099-B!F2 (wash-sale gain)   = {ws['F2'].value}")
    print(f"  1099-B!F3 (no-wash gain)     = {ws['F3'].value}")
    print(f"  1099-B!C{total_row} (SUM total)      = {proceeds_total.value}")
    print(f"  K-1 (1065)!D2 (row TOTAL)    = {ws2['D2'].value}")
    print(f"  header A1 font/fill          = white on #{HEADER_FILL_RGB}")
    print(f"  input C2 font                = blue (#{INPUT_FONT_RGB})")


def main() -> None:
    parser = argparse.ArgumentParser(description="firm-stack tax-workpapers scaffolding builder")
    parser.add_argument("--demo", action="store_true", help="build a small example workbook")
    parser.add_argument(
        "--out",
        default=os.path.join(os.environ.get("TMPDIR", "/tmp"), "tax_workpapers_demo.xlsx"),
        help="output path for --demo (default: a temp path, NOT the skill dir)",
    )
    args = parser.parse_args()

    if args.demo:
        path = build_demo(args.out)
        print(f"Demo workbook written to: {path}")
        _verify(path)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
