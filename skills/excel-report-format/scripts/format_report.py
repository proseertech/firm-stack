#!/usr/bin/env python3
"""
format_report.py — Apply the firm's standard Excel formatting to a workbook.

Part of the firm-stack `excel-report-format` skill. This script implements the
"Firm Formatting Reference" from SKILL.md programmatically: fonts, the firm color
palette, number formats by column type, header/total/subtotal styling, borders,
freeze panes, and print setup.

CORE SAFETY CONSTRAINT (from the skill):
    This is a PRESENTATION pass only. It restyles cells and NEVER changes a
    cell's value or formula. Every write this script performs targets a *style*
    attribute (font, fill, border, alignment, number_format) or a sheet-level
    layout attribute (column width, freeze panes, print setup). It never assigns
    to `cell.value`. See `assert_values_unchanged()` for the guardrail that
    verifies this invariant after formatting.

USAGE:
    python format_report.py input.xlsx [output.xlsx]

    If output.xlsx is omitted, the result is written to
    "<input>-formatted.xlsx" so the original is never overwritten in place.

DEPENDENCIES:
    openpyxl (tested against 3.1.5). No other external dependencies.

TUNING:
    Row/column detection is heuristic and layout-agnostic — it does NOT hard-code
    any single client's spreadsheet. The heuristics are exposed as the well-
    commented constants and helper functions in the "TUNABLE PARAMETERS" section
    below. Adjust those (keyword lists, header-row assumptions, number-format
    column matching) rather than editing the styling engine.
"""

import argparse
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# FIRM FORMATTING REFERENCE — palette (hex codes match SKILL.md EXACTLY)
# ---------------------------------------------------------------------------
# openpyxl expects 6- or 8-digit hex WITHOUT the leading "#". We keep the "#"
# form here as named constants so they read identically to the skill's
# reference table, and strip it at the point of use via `_argb()`.

SECTION_HEADER_FILL = "#1F4E79"   # dark blue  — section headers
COLUMN_HEADER_FILL  = "#D6E4F0"   # soft blue  — column headers
TOTAL_ROW_FILL      = "#E2EFDA"   # soft green — total rows
SUBTOTAL_ROW_FILL   = "#F2F2F2"   # light gray — subtotal rows
FLAGGED_FILL        = "#FFF2CC"   # soft yellow — flagged / callout items

SECTION_HEADER_FONT_COLOR = "#FFFFFF"  # white — section header text
NEGATIVE_FONT_COLOR       = "#C00000"  # red   — negative values
NOTE_FONT_COLOR           = "#666666"  # gray  — notes / comments

# ---------------------------------------------------------------------------
# FIRM FORMATTING REFERENCE — typography
# ---------------------------------------------------------------------------
FONT_NAME = "Calibri"

FONT_SIZE_SECTION_HEADER = 11  # white bold section headers
FONT_SIZE_COLUMN_HEADER  = 10  # bold column headers
FONT_SIZE_DATA           = 10  # normal data rows
FONT_SIZE_NOTE           = 9   # gray italic notes / comments

# ---------------------------------------------------------------------------
# FIRM FORMATTING REFERENCE — number formats
# ---------------------------------------------------------------------------
FMT_CURRENCY            = "$#,##0"            # default currency (no cents)
FMT_CURRENCY_CENTS      = "$#,##0.00"         # when cents matter
FMT_CURRENCY_NEGATIVES  = "$#,##0;($#,##0)"   # negatives in parentheses
FMT_PERCENT             = "0.0%"              # percentages
FMT_INTEGER             = "#,##0"             # integers / counts
FMT_DATE                = "m/d/yyyy"          # dates

# ---------------------------------------------------------------------------
# TUNABLE PARAMETERS — heuristics for detecting row/column roles.
# These are intentionally generic. Adjust for a firm/client that departs from
# the reference layout; do NOT hard-code a specific spreadsheet's structure.
# ---------------------------------------------------------------------------

# A row is treated as a TOTAL row if any of its cells' text starts with one of
# these keywords (case-insensitive). "Total" is the canonical marker.
TOTAL_KEYWORDS = ("total", "grand total", "net ", "net income", "net loss")

# A row is treated as a SUBTOTAL row if any cell's text starts with one of these.
# Checked before TOTAL so "subtotal" isn't swallowed by a loose "total" match.
SUBTOTAL_KEYWORDS = ("subtotal", "sub-total", "sub total")

# A row is treated as a NOTE row if any cell's text starts with one of these.
NOTE_KEYWORDS = ("note:", "notes:", "note ", "footnote", "*")

# Treat the first non-empty row of each sheet as the COLUMN HEADER row. Set to
# False for sheets whose first row is a title/section header rather than column
# labels; you can also override per-sheet via header_row= in format_sheet().
FIRST_ROW_IS_HEADER = True

# If a row's cells are (mostly) bold already, treat it as a structural row
# (header/total/subtotal) even without a keyword. This catches export layouts
# that bold their summary rows but don't label them "Total".
USE_BOLD_AS_STRUCTURAL_HINT = True

# Use parentheses form ($#,##0;($#,##0)) for currency columns so negatives show
# as (1,234). Set False to use a plain "$#,##0" with red font handling negatives.
CURRENCY_NEGATIVES_IN_PARENS = True


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _argb(hex_color):
    """Convert a '#RRGGBB' constant to openpyxl's 'FFRRGGBB' ARGB form."""
    h = hex_color.lstrip("#")
    return ("FF" + h).upper()


def _fill(hex_color):
    argb = _argb(hex_color)
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _font(size, bold=False, italic=False, color=None):
    kwargs = {"name": FONT_NAME, "size": size, "bold": bold, "italic": italic}
    if color:
        kwargs["color"] = _argb(color)
    return Font(**kwargs)


# Precomputed borders per the reference:
#   total row    -> top=medium, bottom=double
#   subtotal row -> bottom=thin
_TOTAL_BORDER = Border(top=Side(style="medium"), bottom=Side(style="double"))
_SUBTOTAL_BORDER = Border(bottom=Side(style="thin"))


# ---------------------------------------------------------------------------
# Row / column role detection (heuristic, tunable)
# ---------------------------------------------------------------------------
def _row_text_cells(ws, row_idx):
    """Return the string values in a row (lowercased, stripped), skipping blanks."""
    out = []
    for cell in ws[row_idx]:
        if isinstance(cell.value, str) and cell.value.strip():
            out.append(cell.value.strip().lower())
    return out


def _starts_with_any(texts, keywords):
    return any(t.startswith(kw) for t in texts for kw in keywords)


def _row_is_mostly_bold(ws, row_idx):
    """True if the populated cells in a row are predominantly bold."""
    bold = 0
    populated = 0
    for cell in ws[row_idx]:
        if cell.value is not None and cell.value != "":
            populated += 1
            if cell.font and cell.font.bold:
                bold += 1
    return populated > 0 and bold >= max(1, populated // 2)


def classify_row(ws, row_idx, header_row):
    """
    Classify a row as one of: 'header', 'subtotal', 'total', 'note', or 'data'.
    Heuristic and tunable via the TUNABLE PARAMETERS constants above.
    """
    if row_idx == header_row:
        return "header"

    texts = _row_text_cells(ws, row_idx)

    if _starts_with_any(texts, NOTE_KEYWORDS):
        return "note"
    # Subtotal checked before total so "subtotal" isn't caught by "total".
    if _starts_with_any(texts, SUBTOTAL_KEYWORDS):
        return "subtotal"
    if _starts_with_any(texts, TOTAL_KEYWORDS):
        return "total"

    # Bold-but-unlabeled rows: treat as subtotal (a conservative structural
    # style) so they read as summary rows without over-claiming "total".
    if USE_BOLD_AS_STRUCTURAL_HINT and _row_is_mostly_bold(ws, row_idx):
        return "subtotal"

    return "data"


def infer_number_format(ws, col_idx, header_row, max_row):
    """
    Infer a number format for a data column by inspecting its header label and
    its data cells. Layout-agnostic: keys off the header text and the cell's
    stored type/format, never a fixed column position.
    Returns a number-format string, or None to leave the column's format alone.
    """
    header_cell = ws.cell(row=header_row, column=col_idx)
    header = str(header_cell.value).strip().lower() if header_cell.value else ""

    # Percentage columns by header keyword or existing % format.
    if "%" in header or "percent" in header or "rate" in header or "margin" in header:
        return FMT_PERCENT

    # Date columns by header keyword or datetime-typed cells.
    if "date" in header:
        return FMT_DATE

    # Scan the column's data cells to decide numeric vs date vs percent.
    saw_number = False
    saw_datetime = False
    saw_fraction_pct = False
    for r in range(header_row + 1, max_row + 1):
        c = ws.cell(row=r, column=col_idx)
        v = c.value
        if v is None:
            continue
        # Respect an existing percent format on the cell.
        if c.number_format and "%" in c.number_format:
            saw_fraction_pct = True
        # openpyxl surfaces dates as datetime objects.
        if hasattr(v, "year") and hasattr(v, "month"):
            saw_datetime = True
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            saw_number = True

    if saw_datetime:
        return FMT_DATE
    if saw_fraction_pct:
        return FMT_PERCENT

    if saw_number:
        # Amount/currency columns by header keyword -> currency; otherwise
        # counts/integers. This is a heuristic default; tune keywords as needed.
        currency_words = ("amount", "balance", "debit", "credit", "$",
                          "income", "expense", "cost", "value", "total",
                          "revenue", "sales", "price", "fee")
        if any(w in header for w in currency_words):
            return FMT_CURRENCY_NEGATIVES if CURRENCY_NEGATIVES_IN_PARENS else FMT_CURRENCY
        # Default numeric columns to currency-with-parens for a TB-style report;
        # counts (qty/units/#) get the integer format.
        count_words = ("qty", "quantity", "count", "units", "#", "number", "id")
        if any(w in header for w in count_words):
            return FMT_INTEGER
        return FMT_CURRENCY_NEGATIVES if CURRENCY_NEGATIVES_IN_PARENS else FMT_CURRENCY

    return None


# ---------------------------------------------------------------------------
# Per-sheet formatting
# ---------------------------------------------------------------------------
def _find_header_row(ws):
    """Return the first non-empty row index, or 1 if the sheet is empty."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                return cell.row
    return 1


def format_sheet(ws, header_row=None, client_name=None):
    """Apply the firm formatting reference to a single worksheet (styles only)."""
    if ws.max_row == 0 or ws.max_column == 0:
        return

    if header_row is None:
        header_row = _find_header_row(ws) if FIRST_ROW_IS_HEADER else -1

    max_row = ws.max_row
    max_col = ws.max_column

    # Pre-infer a number format per column (based on header + data inspection).
    col_formats = {
        c: infer_number_format(ws, c, header_row, max_row)
        for c in range(1, max_col + 1)
    }

    for row_idx in range(1, max_row + 1):
        role = classify_row(ws, row_idx, header_row)

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # --- Fonts / fills / borders per row role ------------------------
            if role == "header":
                # Column-header row: soft-blue fill, bold 10pt.
                cell.fill = _fill(COLUMN_HEADER_FILL)
                cell.font = _font(FONT_SIZE_COLUMN_HEADER, bold=True)
            elif role == "total":
                cell.fill = _fill(TOTAL_ROW_FILL)
                cell.font = _font(FONT_SIZE_DATA, bold=True)
                cell.border = _TOTAL_BORDER
            elif role == "subtotal":
                cell.fill = _fill(SUBTOTAL_ROW_FILL)
                cell.font = _font(FONT_SIZE_DATA, bold=True)
                cell.border = _SUBTOTAL_BORDER
            elif role == "note":
                cell.font = _font(FONT_SIZE_NOTE, italic=True, color=NOTE_FONT_COLOR)
            else:  # data row — white, normal 10pt
                cell.font = _font(FONT_SIZE_DATA)

            # --- Number formats (data/total/subtotal numeric cells) ----------
            # Never touch the header row's format. Apply the inferred column
            # format to numeric cells. This changes DISPLAY only, not the value.
            if role != "header":
                fmt = col_formats.get(col_idx)
                if fmt and isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = fmt
                elif fmt == FMT_DATE and hasattr(cell.value, "year"):
                    cell.number_format = FMT_DATE

            # --- Negative value emphasis -------------------------------------
            # Red font for negative numbers on data rows (reference rule).
            if (role == "data"
                    and isinstance(cell.value, (int, float))
                    and not isinstance(cell.value, bool)
                    and cell.value < 0):
                cell.font = _font(FONT_SIZE_DATA, color=NEGATIVE_FONT_COLOR)

    # --- Column widths: auto-fit to widest cell (no truncation) -------------
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row_idx in range(1, max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                longest = max(longest, len(str(v)))
        # Pad a little; clamp so one long note doesn't create a giant column.
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 60)

    # --- Freeze panes: freeze the header row (row below the header) ----------
    if header_row and header_row >= 1:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # --- Print setup ---------------------------------------------------------
    # Landscape + fit-to-width for wide reports; portrait otherwise.
    ws.page_setup.orientation = (
        "landscape" if max_col > 6 else "portrait"
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:%s%d" % (get_column_letter(max_col), max_row)
    # Header/footer: client name (left) and date (right) if a name was provided.
    if client_name:
        ws.oddHeader.left.text = client_name
    ws.oddHeader.right.text = "&D"  # Excel's built-in current-date field
    ws.print_title_rows = "%d:%d" % (header_row, header_row) if header_row >= 1 else None


# ---------------------------------------------------------------------------
# Value-integrity guardrail
# ---------------------------------------------------------------------------
def snapshot_values(wb):
    """Capture every cell's value keyed by (sheet, row, col) — the safety baseline."""
    snap = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    snap[(ws.title, cell.row, cell.column)] = cell.value
    return snap


def assert_values_unchanged(before, wb):
    """
    Raise if any cell value or formula changed. This enforces the skill's core
    safety constraint: formatting must never alter what the workbook computes.
    """
    after = snapshot_values(wb)
    if before != after:
        diffs = []
        keys = set(before) | set(after)
        for k in sorted(keys):
            b = before.get(k)
            a = after.get(k)
            if b != a:
                diffs.append("  %s: %r -> %r" % (k, b, a))
        raise RuntimeError(
            "SAFETY VIOLATION: cell values changed during formatting:\n"
            + "\n".join(diffs)
        )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Apply the firm's standard Excel formatting to a workbook "
                    "(styles only — never changes values or formulas)."
    )
    parser.add_argument("input", help="Path to the input .xlsx workbook")
    parser.add_argument("output", nargs="?", default=None,
                        help="Output path (default: <input>-formatted.xlsx)")
    parser.add_argument("--client-name", default=None,
                        help="Client name for the print header (optional)")
    args = parser.parse_args(argv)

    if not os.path.isfile(args.input):
        parser.error("input file not found: %s" % args.input)

    output = args.output
    if output is None:
        base, ext = os.path.splitext(args.input)
        output = base + "-formatted" + (ext or ".xlsx")

    # Load with formulas preserved (data_only=False keeps formula strings intact).
    wb = openpyxl.load_workbook(args.input, data_only=False)

    before = snapshot_values(wb)

    for ws in wb.worksheets:
        format_sheet(ws, client_name=args.client_name)

    # Enforce the core safety constraint before writing anything out.
    assert_values_unchanged(before, wb)

    wb.save(output)
    print("Formatted workbook written to: %s" % output)
    print("Sheets formatted: %s" % ", ".join(ws.title for ws in wb.worksheets))
    return 0


if __name__ == "__main__":
    sys.exit(main())
